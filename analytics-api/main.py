from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from google.cloud import bigquery
from datetime import date, timedelta
from concurrent.futures import ThreadPoolExecutor
import openpyxl
import time
import os

# Simple in-memory cache — keyed by (hub_id, from_date, to_date), TTL 5 min
_HUB_CACHE: dict = {}
_CACHE_TTL = 300

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_methods=["GET"], allow_headers=["*"])

PROJECT = "schnell-home-automation"
client  = bigquery.Client(project=PROJECT)

# ── Dock data — loaded once from xlsx at startup ──────────────────────────────
_DOCK_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "dock_data.xlsx")

def _load_dock_rows():
    wb = openpyxl.load_workbook(_DOCK_XLSX)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, r))
        row["total_action_count"] = int(row["total_action_count"] or 0)
        row["success_count"]      = int(row["success_count"]      or 0)
        row["failure_count"]      = int(row["failure_count"]       or 0)
        rows.append(row)
    return rows

_DOCK_ROWS = _load_dock_rows()

def _dock_rows_for_hub(hub_id: str, from_date: str, to_date: str):
    """Return xlsx rows for this hub within the date window."""
    dock_ids = {r["dock_id"] for r in q(f"""
        SELECT DISTINCT dock_id
        FROM `{PROJECT}.schnell_analytics.ha_logs`
        WHERE hub_id = @hub_id AND dock_id IS NOT NULL
    """, [bigquery.ScalarQueryParameter("hub_id", "STRING", hub_id)])}
    return [
        r for r in _DOCK_ROWS
        if r["dock_id"] in dock_ids and from_date <= r["date"] <= to_date
    ]

def _build_dock_stats(rows):
    """Aggregate xlsx rows into per-dock stats with docklets[] sub-array."""
    from collections import defaultdict
    by_dock = defaultdict(lambda: {"total": 0, "success": 0, "failure": 0, "docklets": {}})
    for r in rows:
        dock_id    = r["dock_id"]
        docklet_id = r["docklet_id"]
        by_dock[dock_id]["total"]   += r["total_action_count"]
        by_dock[dock_id]["success"] += r["success_count"]
        by_dock[dock_id]["failure"] += r["failure_count"]
        if docklet_id not in by_dock[dock_id]["docklets"]:
            by_dock[dock_id]["docklets"][docklet_id] = {
                "total": 0, "success": 0, "failure": 0,
                "actions": defaultdict(lambda: {"total": 0, "success": 0, "failure": 0})
            }
        dk = by_dock[dock_id]["docklets"][docklet_id]
        dk["total"]   += r["total_action_count"]
        dk["success"] += r["success_count"]
        dk["failure"] += r["failure_count"]
        a = dk["actions"][r["action"]]
        a["total"]   += r["total_action_count"]
        a["success"] += r["success_count"]
        a["failure"] += r["failure_count"]
    stats = []
    for dock_id, d in by_dock.items():
        rel = round(100 * d["success"] / d["total"], 2) if d["total"] else 0
        docklets = []
        for docklet_id, dk in d["docklets"].items():
            dk_rel = round(100 * dk["success"] / dk["total"], 2) if dk["total"] else 0
            actions = [
                {"action": act, **v,
                 "rel": round(100 * v["success"] / v["total"], 2) if v["total"] else 0}
                for act, v in dk["actions"].items()
            ]
            docklets.append({
                "docklet_id": docklet_id,
                "total":      dk["total"],
                "success":    dk["success"],
                "failure":    dk["failure"],
                "rel":        dk_rel,
                "actions":    actions,
            })
        stats.append({
            "dock_id":  dock_id,
            "total":    d["total"],
            "success":  d["success"],
            "failure":  d["failure"],
            "rel":      rel,
            "docklets": docklets,
        })
    return stats

def _build_dock_usage(rows):
    """Aggregate xlsx rows into dock_usage summary."""
    from collections import defaultdict
    by_action  = defaultdict(int)
    by_docklet = defaultdict(int)
    by_date    = defaultdict(lambda: {"total": 0, "success": 0, "failure": 0})
    for r in rows:
        by_action[r["action"]]    += r["total_action_count"]
        by_docklet[r["docklet_id"]] += r["total_action_count"]
        by_date[r["date"]]["total"]   += r["total_action_count"]
        by_date[r["date"]]["success"] += r["success_count"]
        by_date[r["date"]]["failure"] += r["failure_count"]
    total = sum(by_action.values())
    daily = [
        {"date": d, "day": rows[[r["date"] for r in rows].index(d)]["day_of_week"],
         **v, "rel": round(100 * v["success"] / v["total"], 2) if v["total"] else 0}
        for d, v in sorted(by_date.items())
    ]
    return {
        "total":      total,
        "by_action":  dict(by_action),
        "by_docklet": dict(by_docklet),
        "daily":      daily,
    }

def q(sql, params):
    job = client.query(sql, job_config=bigquery.QueryJobConfig(
        query_parameters=params))
    return [dict(row) for row in job.result()]

def hp(hub_id, from_date, to_date):
    return [
        bigquery.ScalarQueryParameter("hub_id",    "STRING", hub_id),
        bigquery.ScalarQueryParameter("from_date", "DATE",   from_date),
        bigquery.ScalarQueryParameter("to_date",   "DATE",   to_date),
    ]

def uc_src(use_case):
    # maps use_case label → src value that dashboard srcPred() understands
    uc = (use_case or "").lower()
    if "docklet" in uc or "dock" in uc: return "docklet"
    if "remote" in uc:                  return "app_remote"
    if "observ" in uc:                  return "direct_thread"
    return "app"


@app.get("/api/hubs")
def list_hubs():
    rows = q("""
        SELECT DISTINCT hub_id
        FROM `schnell-home-automation.schnell_analytics.app_logs`
        ORDER BY hub_id
    """, [])
    return {"hubs": [r["hub_id"] for r in rows]}


@app.get("/api/hub/{hub_id}")
def hub_detail(hub_id: str,
               from_date: str = Query(default=None),
               to_date:   str = Query(default=None)):
    if not to_date:
        to_date = date.today().isoformat()
    if not from_date:
        from_date = (date.today() - timedelta(days=30)).isoformat()
    from_dt = date.fromisoformat(from_date)
    to_dt   = date.fromisoformat(to_date)
    if (to_dt - from_dt).days > 90:
        from_dt   = to_dt - timedelta(days=90)
        from_date = from_dt.isoformat()
    days_count = (to_dt - from_dt).days + 1
    p  = hp(hub_id, from_date, to_date)
    AL = "`schnell-home-automation.schnell_analytics.app_logs`"
    HL = "`schnell-home-automation.schnell_analytics.ha_logs`"

    # Return cached result if still fresh
    _ck = f"{hub_id}:{from_date}:{to_date}"
    if _ck in _HUB_CACHE:
        _ts, _data = _HUB_CACHE[_ck]
        if time.time() - _ts < _CACHE_TTL:
            return _data

    # ── Fire all 21 queries in parallel ──────────────────────────────────
    with ThreadPoolExecutor(max_workers=15) as ex:
        f_kpi      = ex.submit(q, f"""
            SELECT COUNT(*) AS total, COUNTIF(success) AS success,
                   ROUND(100*COUNTIF(success)/COUNT(*),2) AS reliability,
                   COUNT(DISTINCT entity_id) AS total_devices
            FROM {AL} WHERE hub_id=@hub_id
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date""", p)
        f_le       = ex.submit(q, f"""
            SELECT ROUND(AVG(latency_ms)) AS avg,
                   APPROX_QUANTILES(latency_ms,100)[OFFSET(50)] AS p50,
                   APPROX_QUANTILES(latency_ms,100)[OFFSET(95)] AS p95
            FROM {AL} WHERE hub_id=@hub_id AND latency_ms IS NOT NULL
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date""", p)
        f_le_ev    = ex.submit(q, f"""
            SELECT event_timestamp AS ts, entity_id AS dev, friendly_name,
                   use_case AS uc, latency_ms AS lat, room, network_type AS net,
                   tap_ts AS tap, command_sent_ts AS cmd_sent,
                   rest_response_ts AS rest_resp, ws_confirmation_ts AS ws_conf,
                   trigger_method AS src, success, failure_reason
            FROM {AL} WHERE hub_id=@hub_id AND latency_ms IS NOT NULL
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            ORDER BY event_timestamp DESC LIMIT 50""", p)
        f_hs       = ex.submit(q, f"""
            SELECT ROUND(AVG(ha_processing_latency_ms)) AS avg,
                   APPROX_QUANTILES(ha_processing_latency_ms,100)[OFFSET(50)] AS p50,
                   APPROX_QUANTILES(ha_processing_latency_ms,100)[OFFSET(95)] AS p95
            FROM {HL} WHERE hub_id=@hub_id AND ha_processing_latency_ms IS NOT NULL
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date""", p)
        f_hs_ev    = ex.submit(q, f"""
            SELECT event_timestamp AS ts, entity_id AS dev, friendly_name,
                   ha_event_type AS uc, ha_processing_latency_ms AS lat, room,
                   matter_command_ts AS matter_ts, snap_state_change_ts AS snap_ts
            FROM {HL} WHERE hub_id=@hub_id AND ha_processing_latency_ms IS NOT NULL
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            ORDER BY event_timestamp DESC LIMIT 50""", p)
        f_per_uc   = ex.submit(q, f"""
            SELECT use_case, ROUND(AVG(latency_ms)) AS avg,
                   APPROX_QUANTILES(latency_ms,100)[OFFSET(50)] AS p50,
                   APPROX_QUANTILES(latency_ms,100)[OFFSET(95)] AS p95,
                   ROUND(STDDEV(latency_ms)) AS stddev,
                   COUNT(*) AS count, COUNTIF(success) AS success
            FROM {AL} WHERE hub_id=@hub_id AND latency_ms IS NOT NULL
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            GROUP BY use_case""", p)
        f_per_uc_ev = ex.submit(q, f"""
            SELECT use_case, event_timestamp AS ts, entity_id AS dev, friendly_name,
                   latency_ms AS lat, trigger_method AS src, room
            FROM {AL} WHERE hub_id=@hub_id AND latency_ms IS NOT NULL
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            ORDER BY event_timestamp DESC LIMIT 200""", p)
        f_uc_bkt   = ex.submit(q, f"""
            SELECT use_case,
                   CASE WHEN latency_ms<500  THEN '<500ms'
                        WHEN latency_ms<1000 THEN '500-1000ms'
                        WHEN latency_ms<2000 THEN '1-2s'
                        WHEN latency_ms<5000 THEN '2-5s'
                        ELSE '>5s' END AS bucket,
                   COUNT(*) AS cnt
            FROM {AL} WHERE hub_id=@hub_id AND latency_ms IS NOT NULL
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            GROUP BY use_case, bucket""", p)
        f_bcount   = ex.submit(q, f"""
            SELECT CASE WHEN latency_ms<500  THEN '<500ms'
                        WHEN latency_ms<1000 THEN '500-1000ms'
                        WHEN latency_ms<2000 THEN '1-2s'
                        WHEN latency_ms<5000 THEN '2-5s'
                        ELSE '>5s' END AS bucket,
                   COUNT(*) AS cnt
            FROM {AL} WHERE hub_id=@hub_id AND latency_ms IS NOT NULL
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            GROUP BY bucket ORDER BY MIN(latency_ms)""", p)
        f_bk       = ex.submit(q, f"""
            SELECT CASE WHEN latency_ms<500  THEN '<500ms'
                        WHEN latency_ms<1000 THEN '500-1000ms'
                        WHEN latency_ms<2000 THEN '1-2s'
                        WHEN latency_ms<5000 THEN '2-5s'
                        ELSE '>5s' END AS bucket,
                   event_timestamp AS ts, entity_id AS dev, friendly_name,
                   use_case AS uc, latency_ms AS lat, trigger_method AS src,
                   room, success
            FROM {AL} WHERE hub_id=@hub_id AND latency_ms IS NOT NULL
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            LIMIT 500""", p)
        f_daily    = ex.submit(q, f"""
            SELECT date, COUNT(*) AS total,
                   ROUND(100*COUNTIF(success)/COUNT(*),2) AS rel,
                   APPROX_QUANTILES(latency_ms,100 IGNORE NULLS)[OFFSET(50)] AS p50,
                   ROUND(100*COUNTIF(latency_ms<1000)/
                         NULLIF(COUNTIF(latency_ms IS NOT NULL),0),2) AS ns
            FROM {AL} WHERE hub_id=@hub_id
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            GROUP BY date ORDER BY date""", p)
        f_heat     = ex.submit(q, f"""
            SELECT day_of_week, hour, COUNT(*) AS events,
                   COUNTIF(success = false) AS failures,
                   COUNTIF(use_case IN ('Local App Control','Device Bind (App)')) AS app,
                   COUNTIF(use_case='Docklet Press (App)') AS dock,
                   COUNTIF(use_case='Remote App Control')  AS remote,
                   COUNTIF(use_case='Observed Change (App)') AS auto
            FROM {AL} WHERE hub_id=@hub_id
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            GROUP BY day_of_week, hour""", p)
        f_fail     = ex.submit(q, f"""
            SELECT event_timestamp AS ts, use_case AS uc, entity_id AS dev,
                   friendly_name, failure_reason AS reason, room,
                   trigger_method AS src, CAST(latency_ms AS STRING) AS lat,
                   network_type AS net, COALESCE(docklet_id,'') AS dock
            FROM {AL} WHERE hub_id=@hub_id AND success=false
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            ORDER BY event_timestamp DESC LIMIT 100""", p)
        f_src      = ex.submit(q, f"""
            SELECT use_case, COUNT(*) AS total, COUNTIF(success) AS success,
                   COUNTIF(NOT success) AS fail,
                   ROUND(100*COUNTIF(success)/COUNT(*),2) AS rel
            FROM {AL} WHERE hub_id=@hub_id
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            GROUP BY use_case""", p)
        f_ha_cnt   = ex.submit(q, f"""
            SELECT COUNT(*) AS cnt FROM {HL}
            WHERE hub_id=@hub_id
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date""", p)
        f_fbr      = ex.submit(q, f"""
            SELECT failure_reason AS reason, COUNT(*) AS cnt
            FROM {AL} WHERE hub_id=@hub_id AND success=false AND failure_reason IS NOT NULL
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            GROUP BY failure_reason ORDER BY cnt DESC""", p)
        f_fbr_ev   = ex.submit(q, f"""
            SELECT failure_reason AS reason, event_timestamp AS ts,
                   entity_id AS dev, friendly_name, use_case AS uc,
                   trigger_method AS src, room, CAST(latency_ms AS STRING) AS lat
            FROM {AL} WHERE hub_id=@hub_id AND success=false AND failure_reason IS NOT NULL
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            ORDER BY event_timestamp DESC LIMIT 300""", p)
        f_fbd      = ex.submit(q, f"""
            SELECT entity_id AS dev, failure_reason AS reason, COUNT(*) AS cnt
            FROM {AL} WHERE hub_id=@hub_id AND success=false
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            GROUP BY entity_id, failure_reason ORDER BY entity_id, cnt DESC""", p)
        f_dev      = ex.submit(q, f"""
            SELECT entity_id AS id, ANY_VALUE(room) AS room,
                   COUNT(*) AS total, COUNTIF(success=true) AS success,
                   ROUND(100*COUNTIF(success=true)/COUNT(*),2) AS rel,
                   APPROX_QUANTILES(latency_ms,100 IGNORE NULLS)[OFFSET(50)] AS p50
            FROM {AL} WHERE hub_id=@hub_id
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date
            GROUP BY entity_id ORDER BY total DESC LIMIT 50""", p)
        f_usage    = ex.submit(q, f"""
            SELECT
                COUNTIF(use_case IN ('Local App Control','Device Bind (App)')) AS app,
                COUNTIF(use_case='Docklet Press (App)') AS docklet,
                COUNTIF(use_case='Remote App Control')  AS remote,
                COUNTIF(use_case='Observed Change (App)' AND IFNULL(device_type, '') != 'scene') AS direct
            FROM {AL} WHERE hub_id=@hub_id
              AND DATE(event_timestamp) BETWEEN @from_date AND @to_date""", p)
        f_dock_ids = ex.submit(q, f"""
            SELECT DISTINCT dock_id FROM {HL}
            WHERE hub_id=@hub_id AND dock_id IS NOT NULL""",
            [bigquery.ScalarQueryParameter("hub_id", "STRING", hub_id)])

    # ── Collect results & post-process (fast, sequential) ────────────────
    kpi = f_kpi.result()[0]

    le     = f_le.result()
    le_kpi = le[0] if le else {"avg":0,"p50":0,"p95":0}
    le_events = f_le_ev.result()
    for e in le_events:
        e["src"] = uc_src(e.get("uc"))

    hs     = f_hs.result()
    hs_kpi = hs[0] if hs else {"avg":0,"p50":0,"p95":0}
    hs_events = f_hs_ev.result()

    per_uc_rows = f_per_uc.result()
    per_uc_ev   = f_per_uc_ev.result()
    per_uc_map  = {}
    for e in per_uc_ev:
        uc = e.pop("use_case")
        e["src"] = uc_src(uc)
        per_uc_map.setdefault(uc, []).append(e)

    uc_bkt_rows = f_uc_bkt.result()
    uc_bkt_map: dict = {}
    for r in uc_bkt_rows:
        uc_bkt_map.setdefault(r["use_case"], {})[r["bucket"]] = r["cnt"]

    per_uc = {r["use_case"]: {"avg":r["avg"],"p50":r["p50"],"p95":r["p95"],
              "stddev": int(r["stddev"]) if r["stddev"] is not None else 0,
              "count":r["count"],"success":r["success"],
              "buckets":uc_bkt_map.get(r["use_case"],{}),
              "events":per_uc_map.get(r["use_case"],[])}
              for r in per_uc_rows}

    buckets = {r["bucket"]: r["cnt"] for r in f_bcount.result()}

    bk_rows = f_bk.result()
    bucket_events: dict = {}
    for r in bk_rows:
        bk = r.pop("bucket")
        r["src"] = uc_src(r.get("uc"))
        bucket_events.setdefault(bk, []).append(r)

    daily = f_daily.result()

    heatmap, heatmap_detail = {}, {}
    for r in f_heat.result():
        k = f"{r['day_of_week']}_{r['hour']}"
        events = r["events"]
        failures = r["failures"]
        fail_rate = round(100.0 * failures / events, 1) if events > 0 else 0
        heatmap[k] = fail_rate
        heatmap_detail[k] = {"events":events, "failures":failures, "app":r["app"],"dock":r["dock"],
                              "remote":r["remote"],"auto":r["auto"]}

    failures = []
    for f in f_fail.result():
        f["src"] = uc_src(f.get("uc"))
        f["lat"] = f["lat"] if f["lat"] else "N/A"
        failures.append(f)

    src_rows = f_src.result()
    src_rel  = {r["use_case"]:{k:v for k,v in r.items() if k!="use_case"} for r in src_rows}
    a_rows   = [r for r in src_rows if "App Control" in (r.get("use_case") or "")]
    d_rows   = [r for r in src_rows if "Docklet"     in (r.get("use_case") or "")]
    at=sum(r["total"] for r in a_rows); as_=sum(r["success"] for r in a_rows)
    dt=sum(r["total"] for r in d_rows); ds =sum(r["success"] for r in d_rows)

    ha_cnt_rows       = f_ha_cnt.result()
    hub_to_snap_count = ha_cnt_rows[0]["cnt"] if ha_cnt_rows else 0

    fbr_ev_map: dict = {}
    for e in f_fbr_ev.result():
        r_ = e.pop("reason")
        fbr_ev_map.setdefault(r_, []).append(e)
    fail_by_reason = {
        r["reason"]: {"count": r["cnt"], "events": fbr_ev_map.get(r["reason"], [])}
        for r in f_fbr.result()
    }

    fail_by_device: dict = {}
    for r in f_fbd.result():
        dev = r["dev"] or "unknown"
        if dev not in fail_by_device:
            fail_by_device[dev] = {"count": 0, "reasons": {}}
        fail_by_device[dev]["reasons"][r["reason"] or "UNKNOWN"] = r["cnt"]
        fail_by_device[dev]["count"] += r["cnt"]

    devices = [
        {"id": r["id"] or "unknown", "room": r["room"] or "—",
         "total": int(r["total"]), "success": int(r["success"]),
         "rel": float(r["rel"]) if r["rel"] is not None else 0.0,
         "p50": int(r["p50"]) if r["p50"] is not None else 0}
        for r in f_dev.result()
    ]

    ur          = (f_usage.result() or [{}])[0]
    app_cnt     = int(ur.get("app",     0) or 0)
    docklet_cnt = int(ur.get("docklet", 0) or 0)
    remote_cnt  = int(ur.get("remote",  0) or 0)
    direct_cnt  = int(ur.get("direct",  0) or 0)
    h_total     = app_cnt + docklet_cnt
    usage = {
        "app": app_cnt, "docklet": docklet_cnt,
        "remote": remote_cnt, "direct": direct_cnt,
        "app_ratio":    round(100 * app_cnt     / h_total, 2) if h_total else 0,
        "dock_ratio":   round(100 * docklet_cnt / h_total, 2) if h_total else 0,
        "scene_per_day": round(direct_cnt / days_count, 2),
    }

    # xlsx rows filtered by date only — xlsx has no hub_id column so we trust
    # the user provided data relevant to this system
    _ = f_dock_ids.result()  # consume future (still needed in parallel pool)
    dock_rows_filtered = [r for r in _DOCK_ROWS
                          if from_date <= r["date"] <= to_date]
    dock_usage_data    = _build_dock_usage(dock_rows_filtered)

    result = {
        "total":       kpi["total"],
        "success":     kpi["success"],
        "reliability": kpi["reliability"],
        "total_devices": kpi["total_devices"],
        "speed": {
            "hub_snap_hub": {**hs_kpi, "events": hs_events},
            "local_e2e":    {**le_kpi, "events": le_events},
            "remote_e2e":   {"avg":0,"p50":0,"p95":0,"events":[]},
            "hub_app":      {"avg": hs_kpi.get("avg") or 0,
                             "p50": hs_kpi.get("p50") or 0,
                             "p95": hs_kpi.get("p95") or 0, "events": []},
            "buckets":       buckets,
            "bucket_events": bucket_events,
            "per_uc":        per_uc,
        },
        "daily":          daily,
        "heatmap":        heatmap,
        "heatmap_detail": heatmap_detail,
        "failures":       failures,
        "reliability_detail": {
            "app_trigger_feedback":  round(100*as_/at,2) if at else 0,
            "dock_trigger_feedback": round(100*ds/dt,2)  if dt else 0,
            "hub_to_app":        round(100*as_/hub_to_snap_count,2) if hub_to_snap_count else 0,
            "app_triggers":      at,
            "app_feedbacks":     as_,
            "dock_triggers":     dt,
            "dock_feedbacks":    ds,
            "hub_to_snap_count": hub_to_snap_count,
            "src_rel":           src_rel,
            "dock_stats":        _build_dock_stats(dock_rows_filtered),
        },
        "dock_usage":     dock_usage_data,  # from xlsx
        "usage":          usage,
        "devices":        devices,
        "fail_by_reason": fail_by_reason,
        "fail_by_device": fail_by_device,
    }
    _HUB_CACHE[_ck] = (time.time(), result)
    return result

_root = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")

@app.get("/")
def serve_dashboard():
    return FileResponse(os.path.join(_root, "dashboard.html"))

# Serve dashboard_app.js and any other static files from Analytics/
app.mount("/", StaticFiles(directory=_root), name="static")